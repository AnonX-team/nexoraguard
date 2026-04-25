"""
NexoraGuard Report Generator
Generates PDF and Excel security reports from latest scan data.

Dependencies:
    pip install fpdf2 openpyxl
"""
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


# ── Unicode sanitizer for fpdf2 (Helvetica font is latin-1 only) ─────────────
_UNICODE_MAP = {
    "\u2014": "-",   # em dash  —
    "\u2013": "-",   # en dash  –
    "\u2018": "'",   # left single quote
    "\u2019": "'",   # right single quote
    "\u201c": '"',   # left double quote
    "\u201d": '"',   # right double quote
    "\u2022": "*",   # bullet
    "\u2026": "...", # ellipsis
    "\u00b0": " deg",# degree
    "\u00b1": "+/-", # plus-minus
    "\u00d7": "x",   # multiplication
    "\u00f7": "/",   # division
    "\u2192": "->",  # arrow right
    "\u2190": "<-",  # arrow left
    "\u2713": "OK",  # check mark
    "\u2717": "X",   # cross mark
    "\u26a0": "!",   # warning
    "\u2665": "<3",  # heart
}

def _s(text) -> str:
    """Sanitize text for fpdf2 latin-1 Helvetica font."""
    t = str(text) if text is not None else ""
    for char, repl in _UNICODE_MAP.items():
        t = t.replace(char, repl)
    # Drop any remaining non-latin-1 characters
    return t.encode("latin-1", errors="replace").decode("latin-1")


# ── PDF Report ────────────────────────────────────────────────────────────────

def generate_pdf_report(analysis: dict, snapshot: dict, alerts: list, integrity: list) -> bytes:
    """
    Generate a comprehensive PDF security report.
    Returns raw bytes of the PDF file.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        raise RuntimeError("fpdf2 not installed. Run: pip install fpdf2")

    class NexoraPDF(FPDF):
        def header(self):
            self.set_fill_color(6, 11, 24)
            self.rect(0, 0, 210, 20, 'F')
            self.set_font("Helvetica", "B", 14)
            self.set_text_color(14, 165, 233)
            self.cell(0, 12, "NexoraGuard - Security Report", align="L", new_x="LMARGIN", new_y="NEXT")
            self.set_draw_color(14, 165, 233)
            self.set_line_width(0.3)
            self.line(10, 20, 200, 20)
            self.ln(4)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(100, 100, 100)
            self.cell(0, 5,
                      f"NexoraGuard v2.0  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Page {self.page_no()}",
                      align="C")

    pdf = NexoraPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ── Risk Banner ──────────────────────────────────────────────────────────
    risk       = analysis.get("overall_risk", "UNKNOWN")
    score      = analysis.get("risk_score", 0)
    is_attack  = analysis.get("is_attack", False)
    timestamp  = analysis.get("timestamp", datetime.now().isoformat())[:19].replace("T", " ")

    risk_colors = {
        "CRITICAL": (239, 68, 68),
        "HIGH":     (249, 115, 22),
        "MEDIUM":   (234, 179, 8),
        "LOW":      (34, 197, 94),
        "SAFE":     (100, 116, 139),
        "UNKNOWN":  (100, 116, 139),
    }
    r, g, b = risk_colors.get(risk, (100, 116, 139))

    pdf.set_fill_color(r, g, b)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 22)
    pdf.rect(10, pdf.get_y(), 190, 22, 'F')
    pdf.set_x(10)
    pdf.cell(130, 22, _s(f"  {risk}  -  Score: {score}/100"), border=0)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(60, 22, _s(f"Attack: {'YES' if is_attack else 'NO'}  |  {timestamp}"), border=0, align="R")
    pdf.ln(26)

    # ── Section helper ───────────────────────────────────────────────────────
    def section(title: str):
        pdf.set_fill_color(12, 20, 38)
        pdf.set_text_color(14, 165, 233)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, _s(f"  {title}"), fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(30, 30, 30)
        pdf.ln(2)

    def kv(label: str, value: str, color=None):
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(50, 6, _s(label + ":"), border=0)
        pdf.set_font("Helvetica", "", 9)
        if color:
            pdf.set_text_color(*color)
        else:
            pdf.set_text_color(20, 20, 20)
        pdf.cell(0, 6, _s(str(value)), border=0, new_x="LMARGIN", new_y="NEXT")

    def table_row(cols: list, widths: list, is_header: bool = False):
        if is_header:
            pdf.set_fill_color(20, 30, 50)
            pdf.set_text_color(14, 165, 233)
            pdf.set_font("Helvetica", "B", 8)
        else:
            pdf.set_fill_color(245, 247, 250)
            pdf.set_text_color(30, 30, 30)
            pdf.set_font("Helvetica", "", 8)
        for col, w in zip(cols, widths):
            pdf.cell(w, 7, _s(str(col))[:40], border="B", fill=is_header)
        pdf.ln()

    # ── 1. System Overview ───────────────────────────────────────────────────
    section("1. System Overview")
    stats    = snapshot.get("system_stats", {})
    net_sum  = snapshot.get("network_summary", {})
    hostname = snapshot.get("hostname", "N/A")

    kv("Hostname",     hostname)
    kv("CPU Usage",    f"{stats.get('cpu_percent', 0)}%")
    kv("RAM Usage",    f"{stats.get('ram_percent', 0)}%  ({stats.get('ram_used_gb', 0)} / {stats.get('ram_total_gb', 0)} GB)")
    kv("Disk Free",    f"{stats.get('disk_free_gb', 0)} GB  ({stats.get('disk_percent', 0)}% used)")
    kv("Processes",    str(snapshot.get("total_processes", 0)))
    kv("Connections",  f"{net_sum.get('total', snapshot.get('total_connections', 0))}  (IPv4: {net_sum.get('ipv4_count', 'N/A')}  IPv6: {net_sum.get('ipv6_count', 'N/A')})")
    pdf.ln(4)

    # ── 2. AI Analysis Summary ───────────────────────────────────────────────
    section("2. AI Analysis")
    ai = analysis.get("ai_analysis") or {}
    summary = _s(ai.get("summary", "No AI analysis available."))
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(40, 40, 40)
    pdf.multi_cell(0, 6, summary)
    pdf.ln(2)

    recs = ai.get("recommendations", [])
    if recs:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(14, 165, 233)
        pdf.cell(0, 6, "Recommendations:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(40, 40, 40)
        for i, rec in enumerate(recs[:5], 1):
            pdf.cell(0, 5, _s(f"  {i}. {rec}"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 3. Rule Alerts ───────────────────────────────────────────────────────
    rule_alerts = analysis.get("rule_alerts", [])
    section(f"3. Rule Alerts  ({len(rule_alerts)} detected)")
    if rule_alerts:
        table_row(["Severity", "Rule", "Message", "Time"],
                  [22, 42, 96, 30], is_header=True)
        sev_colors = {"CRITICAL": (239,68,68), "HIGH": (249,115,22), "MEDIUM": (200,140,0), "LOW": (34,197,94)}
        for alert in rule_alerts[:30]:
            sev = alert.get("severity", "")
            pdf.set_fill_color(245, 247, 250)
            pdf.set_text_color(*sev_colors.get(sev, (40,40,40)))
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(22, 7, _s(sev), border="B", fill=True)
            pdf.set_text_color(40, 40, 40)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(42, 7, _s(alert.get("rule", ""))[:20], border="B")
            pdf.cell(96, 7, _s(alert.get("message", ""))[:50], border="B")
            ts = str(alert.get("timestamp", ""))[:16].replace("T", " ")
            pdf.cell(30, 7, ts, border="B")
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, "  No rule alerts fired.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 4. Alert History ─────────────────────────────────────────────────────
    section(f"4. Alert History  (last {len(alerts[:20])} alerts)")
    if alerts:
        table_row(["Risk", "Score", "Attack", "AI Used", "Time"],
                  [25, 20, 20, 22, 103], is_header=True)
        for a in alerts[:20]:
            pdf.set_fill_color(245, 247, 250)
            pdf.set_text_color(40, 40, 40)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(25, 7, a.get("risk", ""), border="B", fill=True)
            pdf.cell(20, 7, str(a.get("score", "")),  border="B")
            pdf.cell(20, 7, "YES" if a.get("is_attack") else "NO", border="B")
            pdf.cell(22, 7, "YES" if a.get("ai_used") else "NO", border="B")
            ts = str(a.get("timestamp", ""))[:19].replace("T", " ")
            pdf.cell(103, 7, ts, border="B")
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, "  No alert history.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 5. File Integrity ────────────────────────────────────────────────────
    section(f"5. File Integrity  ({len(integrity)} violation(s))")
    if integrity:
        table_row(["Status", "Severity", "Path", "Time"],
                  [22, 22, 110, 36], is_header=True)
        for v in integrity[:20]:
            pdf.set_fill_color(245, 247, 250)
            pdf.set_text_color(239, 68, 68)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(22, 7, _s(v.get("status", "")), border="B", fill=True)
            pdf.set_text_color(40, 40, 40)
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(22, 7, _s(v.get("severity", "")), border="B")
            path_short = _s(v.get("path", ""))
            if len(path_short) > 55:
                path_short = "..." + path_short[-52:]
            pdf.cell(110, 7, path_short, border="B")
            ts = str(v.get("timestamp", ""))[:16].replace("T", " ")
            pdf.cell(36, 7, ts, border="B")
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(34, 197, 94)
        pdf.cell(0, 6, "  All monitored files intact - no violations.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # ── 6. Remediation ───────────────────────────────────────────────────────
    rem = analysis.get("remediation", {})
    steps = rem.get("remediation_steps", [])
    section(f"6. Remediation Plan  ({len(steps)} step(s))")
    verdict = rem.get("threat_verdict", "N/A")
    kv("Verdict", verdict)
    if steps:
        for step in steps[:5]:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(14, 165, 233)
            pdf.cell(0, 5, _s(f"  Step {step.get('step', '?')}: [{step.get('priority', '')}] {step.get('action_type', '').upper()}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(40, 40, 40)
            pdf.cell(0, 5, _s(f"    {step.get('description', '')}"), new_x="LMARGIN", new_y="NEXT")
            if step.get("target"):
                pdf.cell(0, 5, _s(f"    Target: {step.get('target')}"), new_x="LMARGIN", new_y="NEXT")
    else:
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, "  No remediation steps required.", new_x="LMARGIN", new_y="NEXT")

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


# ── Excel Report ──────────────────────────────────────────────────────────────

def generate_excel_report(analysis: dict, snapshot: dict, alerts: list, integrity: list) -> bytes:
    """
    Generate a multi-sheet Excel security report.
    Returns raw bytes of the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Color palette ────────────────────────────────────────────────────────
    DARK_BG   = "060B18"
    ACCENT    = "0EA5E9"
    RED       = "EF4444"
    ORANGE    = "F97316"
    YELLOW    = "EAB308"
    GREEN     = "22C55E"
    HEADER_BG = "0C1426"
    ROW_ALT   = "111D33"

    risk_fill = {
        "CRITICAL": PatternFill("solid", fgColor=RED),
        "HIGH":     PatternFill("solid", fgColor=ORANGE),
        "MEDIUM":   PatternFill("solid", fgColor=YELLOW),
        "LOW":      PatternFill("solid", fgColor=GREEN),
        "SAFE":     PatternFill("solid", fgColor="475569"),
    }

    def header_style(cell, text: str):
        cell.value = text
        cell.font  = Font(bold=True, color=ACCENT, size=10)
        cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=ACCENT))

    def set_col_widths(ws, widths: list):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def title_row(ws, text: str):
        ws["A1"] = text
        ws["A1"].font  = Font(bold=True, color="FFFFFF", size=14)
        ws["A1"].fill  = PatternFill("solid", fgColor=DARK_BG)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    ws_sum.sheet_properties.tabColor = ACCENT

    title_row(ws_sum, "NexoraGuard Security Report — Summary")
    ws_sum.merge_cells("A1:D1")

    risk    = analysis.get("overall_risk", "UNKNOWN")
    score   = analysis.get("risk_score", 0)
    is_atk  = "YES" if analysis.get("is_attack") else "NO"
    ai_used = "YES" if (analysis.get("ai_analysis") or {}).get("ai_used") else "NO"
    ts      = str(analysis.get("timestamp", ""))[:19].replace("T", " ")
    stats   = snapshot.get("system_stats", {})
    net_sum = snapshot.get("network_summary", {})

    rows = [
        ("Generated At",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Scan Timestamp",    ts),
        ("Hostname",          snapshot.get("hostname", "N/A")),
        ("", ""),
        ("RISK LEVEL",        risk),
        ("Risk Score",        f"{score} / 100"),
        ("Active Attack",     is_atk),
        ("AI Analysis Used",  ai_used),
        ("Rule Alerts Fired", str(analysis.get("rule_alert_count", 0))),
        ("", ""),
        ("CPU Usage",         f"{stats.get('cpu_percent', 0)}%"),
        ("RAM Usage",         f"{stats.get('ram_percent', 0)}%"),
        ("RAM Used",          f"{stats.get('ram_used_gb', 0)} / {stats.get('ram_total_gb', 0)} GB"),
        ("Disk Free",         f"{stats.get('disk_free_gb', 0)} GB ({stats.get('disk_percent', 0)}% used)"),
        ("Total Processes",   str(snapshot.get("total_processes", 0))),
        ("Total Connections", str(net_sum.get("total", snapshot.get("total_connections", 0)))),
        ("IPv4 Connections",  str(net_sum.get("ipv4_count", "N/A"))),
        ("IPv6 Connections",  str(net_sum.get("ipv6_count", "N/A"))),
        ("Integrity Violations", str(len(integrity))),
    ]

    for r, (label, value) in enumerate(rows, 3):
        ws_sum.cell(r, 1, label).font = Font(bold=True, color="94A3B8", size=9)
        cell = ws_sum.cell(r, 2, value)
        cell.font = Font(color="F1F5F9", size=9)
        ws_sum.cell(r, 1).fill = PatternFill("solid", fgColor="0C1426")
        ws_sum.cell(r, 2).fill = PatternFill("solid", fgColor="111D33")
        # Highlight risk row
        if label == "RISK LEVEL":
            fill = risk_fill.get(risk, PatternFill("solid", fgColor="475569"))
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)

    set_col_widths(ws_sum, [26, 40, 20, 20])

    # ── Sheet 2: Rule Alerts ─────────────────────────────────────────────────
    ws_rules = wb.create_sheet("Rule Alerts")
    ws_rules.sheet_view.showGridLines = False
    ws_rules.sheet_properties.tabColor = RED

    headers = ["Severity", "Rule", "Message", "PID", "Timestamp"]
    for c, h in enumerate(headers, 1):
        header_style(ws_rules.cell(1, c), h)

    rule_alerts = analysis.get("rule_alerts", [])
    sev_fills = {
        "CRITICAL": PatternFill("solid", fgColor="2D0A0A"),
        "HIGH":     PatternFill("solid", fgColor="2D1A0A"),
        "MEDIUM":   PatternFill("solid", fgColor="2D280A"),
        "LOW":      PatternFill("solid", fgColor="0A2D14"),
    }
    for r, alert in enumerate(rule_alerts, 2):
        sev = alert.get("severity", "")
        row_fill = sev_fills.get(sev, PatternFill("solid", fgColor="111D33"))
        vals = [sev, alert.get("rule",""), alert.get("message",""),
                str(alert.get("pid","")), str(alert.get("timestamp",""))[:19]]
        for c, v in enumerate(vals, 1):
            cell = ws_rules.cell(r, c, v)
            cell.fill = row_fill
            cell.font = Font(color="F1F5F9", size=9,
                             bold=(c == 1))
    set_col_widths(ws_rules, [12, 24, 55, 10, 22])

    # ── Sheet 3: Alert History ───────────────────────────────────────────────
    ws_alerts = wb.create_sheet("Alert History")
    ws_alerts.sheet_view.showGridLines = False
    ws_alerts.sheet_properties.tabColor = ORANGE

    headers = ["Risk Level", "Score", "Attack", "AI Used", "Rule Count", "Summary", "Timestamp"]
    for c, h in enumerate(headers, 1):
        header_style(ws_alerts.cell(1, c), h)

    for r, a in enumerate(alerts[:200], 2):
        ai_an = a.get("ai_analysis") or {}
        row_fill = PatternFill("solid", fgColor="111D33" if r % 2 == 0 else "0C1426")
        vals = [
            a.get("risk", ""),
            str(a.get("score", "")),
            "YES" if a.get("is_attack") else "NO",
            "YES" if ai_an.get("ai_used") else "NO",
            str(a.get("rule_alert_count", "")),
            (ai_an.get("summary") or "")[:80],
            str(a.get("timestamp",""))[:19],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_alerts.cell(r, c, v)
            cell.fill = row_fill
            cell.font = Font(color="F1F5F9", size=9)
    set_col_widths(ws_alerts, [13, 8, 8, 8, 11, 60, 22])

    # ── Sheet 4: File Integrity ───────────────────────────────────────────────
    ws_int = wb.create_sheet("File Integrity")
    ws_int.sheet_view.showGridLines = False
    ws_int.sheet_properties.tabColor = RED

    headers = ["Status", "Severity", "Path", "Built-in", "Baseline Time", "Detected At"]
    for c, h in enumerate(headers, 1):
        header_style(ws_int.cell(1, c), h)

    if integrity:
        for r, v in enumerate(integrity, 2):
            row_fill = PatternFill("solid", fgColor="2D0A0A")
            vals = [
                v.get("status",""), v.get("severity",""), v.get("path",""),
                "YES" if v.get("builtin") else "NO",
                str(v.get("baseline_time",""))[:19],
                str(v.get("timestamp",""))[:19],
            ]
            for c, val in enumerate(vals, 1):
                cell = ws_int.cell(r, c, val)
                cell.fill = row_fill
                cell.font = Font(color="F1F5F9", size=9, bold=(c <= 2))
    else:
        ws_int.cell(2, 1, "All monitored files intact — no violations.").font = Font(color=GREEN, bold=True, size=10)
    set_col_widths(ws_int, [14, 12, 55, 10, 22, 22])

    # ── Sheet 5: Remediation ─────────────────────────────────────────────────
    ws_rem = wb.create_sheet("Remediation")
    ws_rem.sheet_view.showGridLines = False
    ws_rem.sheet_properties.tabColor = GREEN

    rem   = analysis.get("remediation", {})
    steps = rem.get("remediation_steps", [])

    ws_rem.cell(1, 1, "Verdict:").font = Font(bold=True, color=ACCENT, size=10)
    ws_rem.cell(1, 2, rem.get("threat_verdict", "N/A")).font = Font(color="F1F5F9", size=10)

    headers = ["Step", "Priority", "Action Type", "Target", "Description"]
    for c, h in enumerate(headers, 1):
        header_style(ws_rem.cell(3, c), h)

    for r, step in enumerate(steps, 4):
        pri = step.get("priority", "")
        p_fill = {"HIGH": PatternFill("solid", fgColor="2D1A0A"),
                  "MEDIUM": PatternFill("solid", fgColor="2D280A"),
                  "LOW": PatternFill("solid", fgColor="0A2D14")}.get(pri, PatternFill("solid", fgColor="111D33"))
        vals = [str(step.get("step","")), pri, step.get("action_type",""),
                str(step.get("target","") or ""), step.get("description","")]
        for c, v in enumerate(vals, 1):
            cell = ws_rem.cell(r, c, v)
            cell.fill = p_fill
            cell.font = Font(color="F1F5F9", size=9, bold=(c <= 2))
    set_col_widths(ws_rem, [7, 10, 16, 15, 60])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
